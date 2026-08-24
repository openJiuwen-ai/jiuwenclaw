import os
from pathlib import Path

from openjiuwen.core.foundation.tool import Tool, ToolCard

_DATE_KEYWORDS = ("date", "time", "created", "updated", "timestamp")


class DataSummarizer(Tool):
    """数据汇总工具：聚合多源数据，整理分类统计并生成结构化报告。"""

    def __init__(self) -> None:
        super().__init__(
            ToolCard(
                id="data_summarizer",
                name="data_summarizer",
                description=(
                    "数据汇总工具：聚合多源数据，自动整理、分类、统计并生成结构化报告。"
                    "当用户需要汇总数据、整理信息、生成报告摘要时调用。"
                ),
                input_params={
                    "type": "object",
                    "properties": {
                        "data_sources": {
                            "type": "array",
                            "items": {"type": "string"},
                            "description": "数据源文件路径列表（支持JSON/CSV/Excel）",
                        },
                        "structured_data": {
                            "type": "array",
                            "items": {"type": "object"},
                            "description": "直接传入的结构化数据列表",
                        },
                        "summary_type": {
                            "type": "string",
                            "enum": [
                                "overview",
                                "statistical",
                                "categorical",
                                "timeline",
                            ],
                            "description": "汇总类型：概览/统计/分类/时间线",
                        },
                        "output_format": {
                            "type": "string",
                            "enum": ["json", "markdown", "structured"],
                            "description": "输出格式：JSON/Markdown/结构化对象",
                        },
                    },
                    "required": [],
                },
            )
        )

    async def invoke(self, inputs, **kwargs):
        data_sources = inputs.get("data_sources", [])
        structured_data = inputs.get("structured_data", [])
        summary_type = inputs.get("summary_type", "overview")
        output_format = inputs.get("output_format", "structured")

        all_data: list = []

        if structured_data:
            all_data.extend(structured_data)

        for source_path in data_sources:
            if not os.path.isfile(source_path):
                return {
                    "success": False,
                    "error": f"数据源文件不存在: {source_path}",
                }
            try:
                file_data = self._load_data_file(source_path)
                if isinstance(file_data, list):
                    all_data.extend(file_data)
                elif isinstance(file_data, dict):
                    all_data.append(file_data)
            except Exception as e:
                return {
                    "success": False,
                    "error": f"读取数据源失败 {source_path}: {e}",
                }

        if not all_data:
            return {
                "success": False,
                "error": "没有可汇总的数据，请提供 data_sources 或 structured_data",
            }

        try:
            summary = self._summarize(all_data, summary_type)

            if output_format == "json":
                import json

                result_output = json.dumps(
                    summary, ensure_ascii=False, indent=2
                )
            elif output_format == "markdown":
                result_output = self._to_markdown(summary)
            else:
                result_output = summary

            return {
                "success": True,
                "summary_type": summary_type,
                "output_format": output_format,
                "total_records": len(all_data),
                "summary": result_output,
            }
        except Exception as e:
            return {"success": False, "error": f"汇总失败: {e}"}

    async def stream(self, inputs, **kwargs):
        yield await self.invoke(inputs, **kwargs)

    @staticmethod
    def _load_data_file(file_path: str):
        ext = Path(file_path).suffix.lower()
        if ext == ".json":
            import json

            with open(file_path, encoding="utf-8") as f:
                return json.load(f)
        if ext == ".csv":
            import csv

            with open(file_path, encoding="utf-8") as f:
                reader = csv.DictReader(f)
                return list(reader)
        if ext in (".xlsx", ".xls"):
            from openpyxl import load_workbook

            wb = load_workbook(file_path, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                return []
            headers = [
                str(h) if h else f"col_{i}" for i, h in enumerate(rows[0])
            ]
            return [dict(zip(headers, row)) for row in rows[1:]]
        import json

        with open(file_path, encoding="utf-8") as f:
            return json.load(f)

    @staticmethod
    def _summarize(data: list, summary_type: str) -> dict:
        total = len(data)
        summary: dict = {
            "total_records": total,
            "summary_type": summary_type,
        }

        all_keys: set[str] = set()
        for item in data:
            if isinstance(item, dict):
                all_keys.update(item.keys())

        if summary_type == "overview":
            field_stats = {}
            for key in all_keys:
                values = [
                    item.get(key)
                    for item in data
                    if isinstance(item, dict) and key in item
                ]
                non_empty = [
                    v for v in values if v is not None and v != ""
                ]
                field_stats[key] = {
                    "count": len(non_empty),
                    "coverage": f"{len(non_empty)}/{total}",
                }
            summary["fields"] = list(all_keys)
            summary["field_stats"] = field_stats
            summary["sample_records"] = data[:5]

        elif summary_type == "statistical":
            numeric_stats = {}
            for key in all_keys:
                values = []
                for item in data:
                    if isinstance(item, dict) and key in item:
                        try:
                            values.append(float(item[key]))
                        except (ValueError, TypeError):
                            pass
                if values:
                    numeric_stats[key] = {
                        "count": len(values),
                        "sum": sum(values),
                        "mean": sum(values) / len(values),
                        "min": min(values),
                        "max": max(values),
                    }
            summary["numeric_fields"] = numeric_stats

        elif summary_type == "categorical":
            categorical_stats = {}
            for key in all_keys:
                value_counts: dict[str, int] = {}
                for item in data:
                    if isinstance(item, dict) and key in item:
                        val = str(item[key])
                        value_counts[val] = value_counts.get(val, 0) + 1
                if value_counts:
                    categorical_stats[key] = {
                        "unique_values": len(value_counts),
                        "distribution": dict(
                            sorted(
                                value_counts.items(),
                                key=lambda x: -x[1],
                            )[:20]
                        ),
                    }
            summary["categorical_fields"] = categorical_stats

        elif summary_type == "timeline":
            date_fields = []
            for key in all_keys:
                lower_key = key.lower()
                if any(kw in lower_key for kw in _DATE_KEYWORDS):
                    date_fields.append(key)
            if date_fields:
                date_field = date_fields[0]
                timeline_items = []
                for item in data:
                    if isinstance(item, dict) and item.get(date_field):
                        timeline_items.append(item)
                sorted_data = sorted(
                    timeline_items,
                    key=lambda x: str(x.get(date_field, "")),
                )
                summary["date_field"] = date_field
                summary["date_fields_detected"] = date_fields
                summary["sorted_records"] = sorted_data
            else:
                summary["date_fields_detected"] = []
                summary["message"] = "未检测到时间字段"

        return summary

    @staticmethod
    def _to_markdown(summary: dict) -> str:
        lines = ["# 数据汇总报告", ""]
        lines.append(f"- 总记录数: {summary.get('total_records', 0)}")
        lines.append(f"- 汇总类型: {summary.get('summary_type', '')}")
        lines.append("")

        if "fields" in summary:
            lines.append("## 字段概览")
            lines.append("")
            for field in summary["fields"]:
                stats = summary.get("field_stats", {}).get(field, {})
                lines.append(
                    f"- **{field}**: 覆盖率 {stats.get('coverage', 'N/A')}"
                )
            lines.append("")

        if "numeric_fields" in summary:
            lines.append("## 数值统计")
            lines.append("")
            for field, stats in summary["numeric_fields"].items():
                lines.append(
                    f"- **{field}**: 总和={stats['sum']:.2f}, "
                    f"均值={stats['mean']:.2f}, "
                    f"最小={stats['min']}, 最大={stats['max']}"
                )
            lines.append("")

        if "categorical_fields" in summary:
            lines.append("## 分类分布")
            lines.append("")
            for field, stats in summary["categorical_fields"].items():
                lines.append(
                    f"- **{field}** ({stats['unique_values']}个唯一值):"
                )
                for val, count in stats["distribution"].items():
                    lines.append(f"  - {val}: {count}")
            lines.append("")

        if "sample_records" in summary:
            lines.append("## 样本数据")
            lines.append("")
            for i, record in enumerate(summary["sample_records"]):
                lines.append(f"### 记录 {i + 1}")
                if isinstance(record, dict):
                    for k, v in record.items():
                        lines.append(f"- {k}: {v}")
                lines.append("")

        return "\n".join(lines)
