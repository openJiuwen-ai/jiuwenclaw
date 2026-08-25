import os
from pathlib import Path

from openjiuwen.core.foundation.tool import Tool, ToolCard


class ExcelTransformer(Tool):
    """Excel数据变换工具：多表合并(vlookup)、数据透视、筛选去重、CSV编码清洗。"""

    def __init__(self) -> None:
        super().__init__(
            ToolCard(
                id="excel_transformer",
                name="excel_transformer",
                description=(
                    "Excel数据变换工具：多表合并(类vlookup)、数据透视、"
                    "按列筛选/去重、CSV编码清洗(GBK<->UTF8)。"
                    "当用户需要对表格数据进行变换操作时调用。"
                ),
                input_params={
                    "type": "object",
                    "properties": {
                        "operation": {
                            "type": "string",
                            "enum": [
                                "vlookup",
                                "pivot",
                                "filter",
                                "dedupe",
                                "encoding_convert",
                                "sort",
                                "split_text",
                                "merge_text",
                            ],
                            "description": "数据变换操作类型",
                        },
                        "file_path": {
                            "type": "string",
                            "description": "主文件路径（vlookup时为主表）",
                        },
                        "second_file_path": {
                            "type": "string",
                            "description": "vlookup时的第二表路径",
                        },
                        "options": {
                            "type": "object",
                            "description": (
                                "操作特定参数：vlookup(key_column/"
                                "columns_to_add)、pivot(rows/cols/values/"
                                "agg_func)、filter(filter_column/filter_value/"
                                "filter_op)、dedupe(dedupe_columns)、"
                                "encoding_convert(target_encoding/"
                                "source_encoding)"
                            ),
                        },
                        "output_dir": {
                            "type": "string",
                            "description": (
                                "产物输出目录的绝对路径。传当前项目目录；"
                                "用户指定了保存位置时用用户指定的目录。"
                            ),
                        },
                    },
                    "required": ["operation", "file_path", "output_dir"],
                },
            )
        )

    async def invoke(self, inputs, **kwargs):
        operation = inputs.get("operation", "")
        file_path = inputs.get("file_path", "")
        second_file_path = inputs.get("second_file_path", "")
        options = inputs.get("options", {})
        output_dir = inputs.get("output_dir", "")

        if not operation:
            return {"success": False, "error": "缺少 operation 参数"}
        if not file_path or not os.path.isfile(file_path):
            return {"success": False, "error": f"文件不存在: {file_path}"}
        if not output_dir:
            return {
                "success": False,
                "error": "缺少 output_dir：请传入当前项目目录的绝对路径",
            }

        base_dir = Path(output_dir).expanduser()
        base_dir.mkdir(parents=True, exist_ok=True)

        try:
            if operation == "vlookup":
                result = self._vlookup(
                    file_path, second_file_path, options, str(base_dir)
                )
            elif operation == "pivot":
                result = self._pivot(file_path, options, str(base_dir))
            elif operation == "filter":
                result = self._filter(file_path, options, str(base_dir))
            elif operation == "dedupe":
                result = self._dedupe(file_path, options, str(base_dir))
            elif operation == "encoding_convert":
                result = self._encoding_convert(
                    file_path, options, str(base_dir)
                )
            elif operation == "sort":
                result = self._sort_data(
                    file_path, options, str(base_dir)
                )
            elif operation == "split_text":
                result = self._split_text(
                    file_path, options, str(base_dir)
                )
            elif operation == "merge_text":
                result = self._merge_text(
                    file_path, options, str(base_dir)
                )
            else:
                return {
                    "success": False,
                    "error": f"不支持的操作: {operation}",
                }

            if "error" in result:
                return {"success": False, "error": result["error"]}
            return {"success": True, "operation": operation, **result}
        except ImportError as e:
            return {
                "success": False,
                "error": f"依赖库缺失: {e}. 请安装对应依赖后重试。",
            }
        except Exception as e:
            return {"success": False, "error": f"变换失败: {e}"}

    async def stream(self, inputs, **kwargs):
        yield await self.invoke(inputs, **kwargs)

    @staticmethod
    def _load_excel_data(file_path: str) -> list[dict]:
        """Load Excel/CSV as list of dicts."""
        ext = Path(file_path).suffix.lower()
        if ext == ".csv":
            import csv

            with open(file_path, encoding="utf-8") as f:
                reader = csv.DictReader(f)
                return list(reader)
        from openpyxl import load_workbook

        wb = load_workbook(file_path, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [
            str(h) if h else f"col_{i}" for i, h in enumerate(rows[0])
        ]
        return [dict(zip(headers, row)) for row in rows[1:]]

    @staticmethod
    def _save_excel_data(data: list[dict], output_path: str) -> None:
        """Save list of dicts to Excel."""
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "Result"
        if not data:
            wb.save(output_path)
            return
        headers = list(data[0].keys())
        ws.append(headers)
        for row in data:
            ws.append([row.get(h, "") for h in headers])
        wb.save(output_path)

    def _vlookup(
        self,
        main_path: str,
        second_path: str,
        options: dict,
        output_dir: str,
    ) -> dict:
        if not second_path or not os.path.isfile(second_path):
            return {"error": "vlookup需要 second_file_path 参数"}
        key_column = options.get("key_column", "")
        columns_to_add = options.get("columns_to_add", [])
        if not key_column:
            return {"error": "vlookup需要 key_column 参数"}

        main_data = self._load_excel_data(main_path)
        second_data = self._load_excel_data(second_path)

        lookup: dict[str, dict] = {}
        for row in second_data:
            key = str(row.get(key_column, ""))
            if key:
                lookup[key] = row

        matched = 0
        for row in main_data:
            key = str(row.get(key_column, ""))
            if key in lookup:
                matched += 1
                second_row = lookup[key]
                if columns_to_add:
                    for col in columns_to_add:
                        if col in second_row:
                            row[col] = second_row[col]
                else:
                    for col in second_row:
                        if col not in row:
                            row[col] = second_row[col]

        output_path = str(Path(output_dir) / "vlookup_result.xlsx")
        self._save_excel_data(main_data, output_path)
        return {
            "total_rows": len(main_data),
            "matched": matched,
            "unmatched": len(main_data) - matched,
            "path": output_path,
            "exists": os.path.isfile(output_path),
            "size_bytes": (
                os.path.getsize(output_path)
                if os.path.isfile(output_path)
                else 0
            ),
        }

    def _pivot(
        self, file_path: str, options: dict, output_dir: str
    ) -> dict:
        rows_field = options.get("rows", "")
        cols_field = options.get("cols", "")
        values_field = options.get("values", "")
        agg_func = options.get("agg_func", "sum")

        if not rows_field or not values_field:
            return {"error": "pivot需要 rows 和 values 参数"}

        data = self._load_excel_data(file_path)

        pivot_data: dict[tuple, list] = {}
        col_keys: set[str] = set()
        for row in data:
            r_key = str(row.get(rows_field, ""))
            c_key = (
                str(row.get(cols_field, "")) if cols_field else "Value"
            )
            col_keys.add(c_key)
            try:
                val = float(row.get(values_field, 0))
            except (ValueError, TypeError):
                val = 0
            if (r_key, c_key) not in pivot_data:
                pivot_data[(r_key, c_key)] = []
            pivot_data[(r_key, c_key)].append(val)

        sorted_col_keys = sorted(col_keys)
        result_rows: list[dict] = []
        row_keys = sorted({k[0] for k in pivot_data})
        for r_key in row_keys:
            result_row: dict = {rows_field: r_key}
            for c_key in sorted_col_keys:
                vals = pivot_data.get((r_key, c_key), [])
                if not vals:
                    result_row[c_key] = 0
                elif agg_func == "sum":
                    result_row[c_key] = sum(vals)
                elif agg_func == "avg":
                    result_row[c_key] = sum(vals) / len(vals)
                elif agg_func == "count":
                    result_row[c_key] = len(vals)
                elif agg_func == "max":
                    result_row[c_key] = max(vals)
                elif agg_func == "min":
                    result_row[c_key] = min(vals)
                else:
                    result_row[c_key] = sum(vals)
            result_rows.append(result_row)

        output_path = str(Path(output_dir) / "pivot_result.xlsx")
        self._save_excel_data(result_rows, output_path)
        return {
            "total_rows": len(result_rows),
            "pivot_rows": rows_field,
            "pivot_cols": cols_field or "Value",
            "agg_func": agg_func,
            "path": output_path,
            "exists": os.path.isfile(output_path),
            "size_bytes": (
                os.path.getsize(output_path)
                if os.path.isfile(output_path)
                else 0
            ),
        }

    def _filter(
        self, file_path: str, options: dict, output_dir: str
    ) -> dict:
        filter_column = options.get("filter_column", "")
        filter_value = options.get("filter_value", "")
        filter_op = options.get("filter_op", "equals")

        if not filter_column:
            return {"error": "filter需要 filter_column 参数"}

        data = self._load_excel_data(file_path)
        original_count = len(data)

        filtered = []
        for row in data:
            val = str(row.get(filter_column, ""))
            if filter_op == "equals" and val == str(filter_value):
                filtered.append(row)
            elif filter_op == "contains" and str(filter_value) in val:
                filtered.append(row)
            elif filter_op == "not_equals" and val != str(filter_value):
                filtered.append(row)
            elif (
                filter_op == "starts_with"
                and val.startswith(str(filter_value))
            ):
                filtered.append(row)
            elif (
                filter_op == "ends_with"
                and val.endswith(str(filter_value))
            ):
                filtered.append(row)

        output_path = str(Path(output_dir) / "filter_result.xlsx")
        self._save_excel_data(filtered, output_path)
        return {
            "original_count": original_count,
            "filtered_count": len(filtered),
            "filter_column": filter_column,
            "filter_op": filter_op,
            "path": output_path,
            "exists": os.path.isfile(output_path),
            "size_bytes": (
                os.path.getsize(output_path)
                if os.path.isfile(output_path)
                else 0
            ),
        }

    def _dedupe(
        self, file_path: str, options: dict, output_dir: str
    ) -> dict:
        dedupe_columns = options.get("dedupe_columns", [])

        data = self._load_excel_data(file_path)
        original_count = len(data)

        seen: set = set()
        unique: list[dict] = []
        for row in data:
            if not dedupe_columns:
                key = tuple(sorted(row.items()))
            else:
                key = tuple(
                    str(row.get(col, "")) for col in dedupe_columns
                )
            if key not in seen:
                seen.add(key)
                unique.append(row)

        output_path = str(Path(output_dir) / "dedupe_result.xlsx")
        self._save_excel_data(unique, output_path)
        return {
            "original_count": original_count,
            "unique_count": len(unique),
            "duplicates_removed": original_count - len(unique),
            "dedupe_columns": dedupe_columns or "all",
            "path": output_path,
            "exists": os.path.isfile(output_path),
            "size_bytes": (
                os.path.getsize(output_path)
                if os.path.isfile(output_path)
                else 0
            ),
        }

    @staticmethod
    def _encoding_convert(
        file_path: str, options: dict, output_dir: str
    ) -> dict:
        target_encoding = options.get("target_encoding", "utf-8")
        source_encoding = options.get("source_encoding", "auto")

        ext = Path(file_path).suffix.lower()
        if ext != ".csv":
            return {"error": "编码转换仅支持CSV文件"}

        if source_encoding == "auto":
            detected = None
            for enc in ("utf-8", "gbk", "gb2312", "latin-1"):
                try:
                    with open(file_path, encoding=enc) as f:
                        content = f.read()
                    detected = enc
                    break
                except (UnicodeDecodeError, OSError):
                    continue
            if detected is None:
                return {"error": "无法自动检测文件编码"}
            source_encoding = detected
        else:
            with open(file_path, encoding=source_encoding) as f:
                content = f.read()

        stem = Path(file_path).stem
        output_path = str(
            Path(output_dir) / f"{stem}_{target_encoding}.csv"
        )
        with open(output_path, "w", encoding=target_encoding, newline="") as f:
            f.write(content)

        return {
            "source_encoding": source_encoding,
            "target_encoding": target_encoding,
            "path": output_path,
            "exists": os.path.isfile(output_path),
            "size_bytes": (
                os.path.getsize(output_path)
                if os.path.isfile(output_path)
                else 0
            ),
        }

    def _sort_data(
        self, file_path: str, options: dict, output_dir: str
    ) -> dict:
        sort_columns = options.get("sort_columns", [])
        sort_order = options.get("sort_order", "asc")

        if not sort_columns:
            return {"error": "sort需要 sort_columns 参数"}

        data = self._load_excel_data(file_path)
        original_count = len(data)

        reverse = sort_order == "desc"
        data.sort(
            key=lambda row: tuple(
                str(row.get(col, "")) for col in sort_columns
            ),
            reverse=reverse,
        )

        output_path = str(
            Path(output_dir) / "sort_result.xlsx"
        )
        self._save_excel_data(data, output_path)
        return {
            "original_count": original_count,
            "sorted_count": len(data),
            "sort_columns": sort_columns,
            "sort_order": sort_order,
            "path": output_path,
            "exists": os.path.isfile(output_path),
            "size_bytes": (
                os.path.getsize(output_path)
                if os.path.isfile(output_path)
                else 0
            ),
        }

    def _split_text(
        self, file_path: str, options: dict, output_dir: str
    ) -> dict:
        column = options.get("column", "")
        delimiter = options.get("delimiter", ",")
        new_column_prefix = options.get(
            "new_column_prefix", "split_"
        )

        if not column:
            return {"error": "split_text需要 column 参数"}

        data = self._load_excel_data(file_path)

        max_splits = 0
        for row in data:
            val = str(row.get(column, ""))
            parts = val.split(delimiter)
            max_splits = max(max_splits, len(parts))

        for row in data:
            val = str(row.get(column, ""))
            parts = val.split(delimiter)
            for i in range(max_splits):
                row[f"{new_column_prefix}{i + 1}"] = (
                    parts[i].strip()
                    if i < len(parts)
                    else ""
                )

        output_path = str(
            Path(output_dir) / "split_text_result.xlsx"
        )
        self._save_excel_data(data, output_path)
        return {
            "total_rows": len(data),
            "column": column,
            "delimiter": delimiter,
            "new_columns": max_splits,
            "path": output_path,
            "exists": os.path.isfile(output_path),
            "size_bytes": (
                os.path.getsize(output_path)
                if os.path.isfile(output_path)
                else 0
            ),
        }

    def _merge_text(
        self, file_path: str, options: dict, output_dir: str
    ) -> dict:
        columns = options.get("columns", [])
        separator = options.get("separator", " ")
        new_column = options.get("new_column", "merged")

        if not columns:
            return {"error": "merge_text需要 columns 参数"}

        data = self._load_excel_data(file_path)

        for row in data:
            values = [
                str(row.get(col, ""))
                for col in columns
                if row.get(col, "")
            ]
            row[new_column] = separator.join(values)

        output_path = str(
            Path(output_dir) / "merge_text_result.xlsx"
        )
        self._save_excel_data(data, output_path)
        return {
            "total_rows": len(data),
            "columns": columns,
            "separator": separator,
            "new_column": new_column,
            "path": output_path,
            "exists": os.path.isfile(output_path),
            "size_bytes": (
                os.path.getsize(output_path)
                if os.path.isfile(output_path)
                else 0
            ),
        }
