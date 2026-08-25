"""Excel AIGC decorator."""
import zipfile
from xml.etree import ElementTree as ET

from openpyxl import load_workbook
from openpyxl.packaging.custom import StringProperty

from decorators.common import get_aigc_signature, is_aigc_complete, parse_aigc_json


class ExcelAigcDecorator:
    """Excel AIGC decorator - adds hidden AIGC mark to Excel files."""

    def __init__(self):
        self.name = "excel_aigc_decorator"

    def decorate(self, file_path: str, content: str, add_visible_mark: bool = True):
        """Add AIGC mark to Excel file."""
        existing = self._get_aigc_data(file_path)
        if existing and is_aigc_complete(existing):
            print(f"  [Excel] AIGC mark complete, skipping")
            return
        try:
            aigc_signature = get_aigc_signature(content)
            self._add_aigc_mark(file_path, aigc_signature, add_visible_mark)
            print(f"  [Excel] AIGC mark added successfully")
        except Exception as e:
            print(f"  [Excel] Warning: Failed to add AIGC mark: {str(e)}")

    def _get_aigc_data(self, file_path: str) -> dict | None:
        """Read and parse AIGC metadata from Excel custom properties."""
        try:
            with zipfile.ZipFile(file_path, 'r') as zf:
                if 'docProps/custom.xml' not in zf.namelist():
                    return None
                xml_content = zf.read('docProps/custom.xml')
                root = ET.fromstring(xml_content)
                for prop in root.iter():
                    if prop.get('name') == 'AIGC':
                        for child in prop:
                            if child.text:
                                return parse_aigc_json(child.text)
                        return None
        except (zipfile.BadZipFile, ET.ParseError, KeyError):
            return None
        return None

    def _add_aigc_mark(self, file_path: str, signature: str, add_visible_mark: bool = True):
        """Add custom property to Excel file."""
        wb = load_workbook(file_path)
        if add_visible_mark:
            self.add_visible_mark(wb)
        self._add_implicit_mark(wb, signature)
        wb.save(file_path)

    def _add_implicit_mark(self, wb, signature):
        custom_props = wb.custom_doc_props
        if custom_props is not None:
            # Try to find and update existing AIGC property
            for prop in custom_props:
                if hasattr(prop, 'name') and prop.name == 'AIGC':
                    if hasattr(prop, 'value'):
                        prop.value = signature
                    return
            # Not found, append new property
            custom_props.append(StringProperty(name='AIGC', value=signature))

    def add_visible_mark(self, workbook):
        """
        添加显示AI标识

        在每个sheet的最后一行添加"内容由AI生成"
        """
        ai_mark_content = "内容由AI生成"

        for sheet_name in workbook.sheetnames:
            ws = workbook[sheet_name]
            ws.append([ai_mark_content])
